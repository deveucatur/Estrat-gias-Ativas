import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image
from openpyxl import load_workbook
codAcessoNA = ["0000","0001"]
icone = Image.open(f'icone.png')
st.set_page_config(
page_title="Edit UP",
page_icon = icone,
layout="centered")


with st.sidebar:

	add_radio = st.radio("", ("Editar", "Criar Nova Estratégia"))


if add_radio == "Editar":
	st.image(Image.open(f'logo UP.png'))
	st.title("Editar")

	col1, col2 = st.columns((1,2))
	with col1:
	    codA = st.text_input("Código de acesso", type='password')
	
	if codA not in codAcessoNA:
	    with col2:
	        st.text_input("", "Código não autorizado")
	else:
		
		
		df = pd.read_excel("UP2.xlsx", sheet_name = "Página2")
		listDados = df.values.tolist()

		#selection = st.selectbox("selecione a estratégia", df["ESTRATÉGIA"])
		selection = st.selectbox("Selecione a estratégia", [x[0] for x in listDados])
		st.write("---")


		option = [x for x in range(len(listDados)) if listDados[x][0] == selection]

		st.subheader("Destinação")
		dest = st.text_area("Colocar Primeira Letra Maiúscula, separado por vírgula e sem espaço. EX:.(Individual, Pares, Pequeno grupo) ", listDados[option[0]][1])
		
		st.subheader("Breve descrição")
		breDes = st.text_area("Breve descrição", listDados[option[0]][2])
		st.write("---")
		st.subheader("Passo a passo didático")
		ppDid = st.text_area("Passo a passo didático",listDados[option[0]][3])
		st.write("---")
		st.subheader("Nível Cognitivo Potencializado na Taxonomia de Bloom")
		NivCog = st.text_area("Colocar Primeira Letra Maiúscula, separado por vírgula e sem espaço. EX:.(Relembrar, Entender) ",listDados[option[0]][4])
		st.write("---")
		st.subheader("Competências comportamentais que podem ser desenvolvidas")
		compC = st.text_area('Colocar Primeira Letra Maiúscula, separado por vírgula e sem espaço. EX:.(Comunicação, Gestão do tempo, Planejamento, Liderança, Relacionamento)',listDados[option[0]][5])
		st.write("---")
		st.subheader("Referências")
		Ref = st.text_area("Referências", listDados[option[0]][6])
		st.write("---")
		if st.button("Editar"):
		#next_empty_row = len(listDados2) + 2
			# lendo o arquivo excel existente
			workbook = load_workbook("UP2.xlsx")
			sheet = workbook.get_sheet_by_name('Página2')
			# encontrando a próxima linha vazia
			
			# adicionando dados na próxima linha vazia
			sheet.cell(row=option[0]+2, column=2).value = dest
			sheet.cell(row=option[0]+2, column=3).value = breDes
			sheet.cell(row=option[0]+2, column=4).value = ppDid
			sheet.cell(row=option[0]+2, column=5).value = NivCog
			sheet.cell(row=option[0]+2, column=6).value = compC
			sheet.cell(row=option[0]+2, column=7).value = Ref
			# salvando as alterações no arquivo excel
			workbook.save('UP2.xlsx')
			st.info(f"Atualizado")



if add_radio == "Criar Nova Estratégia":
	st.image(Image.open(f'logo UP.png'))
	st.title("Criar Nova Estratégia")
	
	

	col1, col2 = st.columns((1,2))
	with col1:
	    codA = st.text_input("Código de acesso", type='password')
	
	if codA not in codAcessoNA:
	    with col2:
	        st.text_input("", "Código não autorizado")
	else:

		#selection = st.selectbox("selecione a estratégia", df["ESTRATÉGIA"])
		Nome = st.text_input("Nome da Estratégia")
		st.write("---")

		st.subheader("Destinação")
		dest = st.text_area("Colocar Primeira Letra Maiúscula, separado por vírgula e sem espaço. EX:.(Individual, Pares, Pequeno grupo) ")
		
		st.subheader("Breve descrição")
		breDes = st.text_area("Breve descrição")
		st.write("---")
		st.subheader("Passo a passo didático")
		ppDid = st.text_area("Passo a passo didático")
		st.write("---")
		st.subheader("Nível Cognitivo Potencializado na Taxonomia de Bloom")
		NivCog = st.text_area("Colocar Primeira Letra Maiúscula, separado por vírgula e sem espaço. EX:.(Relembrar, Entender) ")
		st.write("---")
		st.subheader("Competências comportamentais que podem ser desenvolvidas")
		compC = st.text_area('Colocar Primeira Letra Maiúscula, separado por vírgula e sem espaço. EX:.(Comunicação, Gestão do tempo, Planejamento, Liderança, Relacionamento)')
		st.write("---")
		st.subheader("Referências")
		Ref = st.text_area("Referências", )
		st.write("---")
		if st.button("Editar"):
			df = pd.read_excel("UP2.xlsx", sheet_name = "Página2")
			listDados = df.values.tolist()

			next_empty_row = len(listDados) + 2

			# lendo o arquivo excel existente
			workbook = load_workbook("UP2.xlsx")
			sheet = workbook.get_sheet_by_name('Página2')
			# encontrando a próxima linha vazia
			
			# adicionando dados na próxima linha vazia
			sheet.cell(row=next_empty_row, column=1).value = Nome
			sheet.cell(row=next_empty_row, column=2).value = dest
			sheet.cell(row=next_empty_row, column=3).value = breDes
			sheet.cell(row=next_empty_row, column=4).value = ppDid
			sheet.cell(row=next_empty_row, column=5).value = NivCog
			sheet.cell(row=next_empty_row, column=6).value = compC
			sheet.cell(row=next_empty_row, column=7).value = Ref
			# salvando as alterações no arquivo excel
			workbook.save('UP2.xlsx')
			st.info(f"Atualizado")











st.write("")
st.write("")
st.write("")
st.write("<center>Todos os direitos reservados © 2023 v0.0.1 - UP - Unindo potencialidades",unsafe_allow_html = True)
